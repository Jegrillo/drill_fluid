import openpyxl
def import_data1(x, y, dat1):
    workbook = openpyxl.load_workbook('data.xlsx')
    sheet = workbook.active
    cell = sheet.cell(row=x, column=y)
    cell.value = dat1
    workbook.save('data.xlsx')

def import_data2(x, niz, verh, davlenie1, davlenie2):
    workbook = openpyxl.load_workbook('intervals.xlsx')
    sheet = workbook.active
    cell = sheet.cell(row=x, column=1)
    cell.value = niz
    cell = sheet.cell(row=x, column=2)
    cell.value = verh
    cell = sheet.cell(row=x, column=3)
    cell.value = davlenie1
    cell = sheet.cell(row=x, column=4)
    cell.value = davlenie2
    workbook.save( 'intervals.xlsx')